from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from tqdm import tqdm
import auth_date
import datetime
import time
import csv
import openpyxl


def teka():
    # Формирование Teka
    file_csv = "teka.csv"
    xpath_manufacturer = "/html/body/main/section/div[2]/div[1]/div[2]/div[3]/ul/li[6]"

    book = openpyxl.open('teka.xlsx', read_only=True)
    sheet = book.active

    with open(f"teka.csv", "w", newline='', encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(
            (
                "Артикул",
                "Наличие"
            )
        )

    for row in tqdm(range(17, sheet.max_row + 1)):
        sku = str(sheet[row][0].value)
        quantity = str(sheet[row][2].value)
        if quantity > '0':
            quantity = 'В наявності'
        with open(f"teka.csv", "a", newline='', encoding="utf-8") as file:
            writer = csv.writer(file)
            writer.writerow(
                (
                    sku,
                    quantity
                )
            )
    print("TEKA - ОК")

    browser_update(file_csv, xpath_manufacturer)

def franke():
    # Формирование Franke
    file_csv = "franke.csv"
    xpath_manufacturer = "/html/body/main/section/div[2]/div[1]/div[2]/div[3]/ul/li[8]/div"

    book = openpyxl.open('franke.xlsx', read_only=True)
    sheet = book.active

    with open(f"franke.csv", "w", newline='', encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(
            (
                "Артикул",
                "Наличие"
            )
        )

    for row in tqdm(range(2, sheet.max_row + 1)):
        sku = sheet[row][1].value
        quantity = str(sheet[row][3].value)
        if quantity == 'наявність по запиту':
            continue
        else:
            if quantity > '0':
                quantity = 'В наявності'
            with open(f"franke.csv", "a", newline='', encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(
                    (
                        sku,
                        quantity
                    )
                )
    print("Franke - ОК")

    browser_update(file_csv, xpath_manufacturer)

def bsh():
    # Формирование BSH
    file_csv = "bsh.csv"
    xpath_manufacturer = "/html/body/main/section/div[2]/div[1]/div[2]/div[3]/ul/li[5]/div"

    book = openpyxl.open('bsh.xlsx', read_only=True)
    sheet = book.active

    with open(f"bsh.csv", "w", newline='', encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(
            (
                "Артикул",
                "Наличие"
            )
        )

    for row in tqdm(range(13, sheet.max_row + 1)):
        try:
            sku = sheet[row][4].value
            quantity = sheet[row][7].value
            quantity = quantity.replace("да", "В наявності")
            if quantity == 'нет':
                continue
            else:
                with open(f"bsh.csv", "a", newline='', encoding="utf-8") as file:
                    writer = csv.writer(file)
                    writer.writerow(
                        (
                            sku,
                            quantity
                        )
                    )
        except Exception as ex:
            print(ex)
    print("BSH - ОК")

    browser_update(file_csv, xpath_manufacturer)

def mirs():
    # Формирование Mirs
    file_csv = "mirs.csv"
    xpath_manufacturer = "/html/body/main/section/div[2]/div[1]/div[2]/div[3]/ul/li[2]/div"

    book = openpyxl.open('mirs.xlsx', read_only=True)
    sheet = book.active
    mirs_all = []

    with open(f"mirs.csv", "w", newline='', encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(
            (
                "Артикул",
                "Наличие"
            )
        )

    for row in tqdm(range(11, sheet.max_row + 1)):
        manufacturer = sheet[row][1].value
        sku = sheet[row][2].value
        quantity = str(sheet[row][9].value)
        if quantity > '0':
            quantity = 'В наявності'
        mirs = []
        if quantity == 'В наявності':
            if manufacturer == 'BLANCO':
                with open(f"mirs.csv", "a", newline='', encoding="utf-8") as file:
                    writer = csv.writer(file)
                    writer.writerow(
                        (
                            sku,
                            quantity
                        )
                    )
            elif manufacturer == 'Falmec':
                with open(f"mirs.csv", "a", newline='', encoding="utf-8") as file:
                    writer = csv.writer(file)
                    writer.writerow(
                        (
                            sku,
                            quantity
                        )
                    )
            elif manufacturer == 'Liebherr':
                with open(f"mirs.csv", "a", newline='', encoding="utf-8") as file:
                    writer = csv.writer(file)
                    writer.writerow(
                        (
                            sku,
                            quantity
                        )
                    )
            elif manufacturer == 'Nivona':
                with open(f"mirs.csv", "a", newline='', encoding="utf-8") as file:
                    writer = csv.writer(file)
                    writer.writerow(
                        (
                            sku,
                            quantity
                        )
                    )
            elif manufacturer == 'Vestel':
                with open(f"mirs.csv", "a", newline='', encoding="utf-8") as file:
                    writer = csv.writer(file)
                    writer.writerow(
                        (
                            sku,
                            quantity
                        )
                    )
            else:
                continue
        else:
            continue
    print("Mirs - ОК")

    browser_update(file_csv, xpath_manufacturer)

def smeg():
    # Формирование Smeg
    file_csv = "smeg.csv"
    xpath_manufacturer = "/html/body/main/section/div[2]/div[1]/div[2]/div[3]/ul/li[4]/div"

    book = openpyxl.open('smeg_mbt.xlsx', read_only=True)
    sheet = book.active

    with open(f"smeg.csv", "w", newline='', encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(
            (
                "Артикул",
                "Наличие"
            )
        )

    for row in tqdm(range(9, sheet.max_row + 1)):
        sku = sheet[row][4].value
        quantity = str(sheet[row][13].value)
        if quantity != '1':
            continue
        else:
            quantity = quantity.replace('1', 'В наявності')
            with open(f"smeg.csv", "a", newline='', encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(
                    (
                        sku,
                        quantity
                    )
                )
    book = openpyxl.open('smeg_kbt.xlsx', read_only=True)
    sheet = book.active
    for row in tqdm(range(9, sheet.max_row + 1)):
        sku = sheet[row][4].value
        quantity = str(sheet[row][13].value)
        if quantity != '1':
            continue
        else:
            quantity = quantity.replace('1', 'В наявності')
            with open(f"smeg.csv", "a", newline='', encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(
                    (
                        sku,
                        quantity
                    )
                )

    print("Smeg - ОК")

    browser_update(file_csv, xpath_manufacturer)

def renklod():
    # Формирование Renklod
    file_csv = "renklod.csv"
    xpath_manufacturer = "/html/body/main/section/div[2]/div[1]/div[2]/div[3]/ul/li[7]/div"

    book = openpyxl.open('renklod.xlsx', read_only=True)
    sheet = book.active

    with open(f"renklod.csv", "w", newline='', encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(
            (
                "Артикул",
                "Наличие"
            )
        )

    for row in tqdm(range(2, sheet.max_row + 1)):
        sku = sheet[row][5].value
        quantity = str(sheet[row][8].value)
        quantity = 'В наявності'
        with open(f"renklod.csv", "a", newline='', encoding="utf-8") as file:
            writer = csv.writer(file)
            writer.writerow(
                (
                    sku,
                    quantity
                )
            )

    print("Renklod - ОК")

    browser_update(file_csv, xpath_manufacturer)

def browser_update(file_csv, xpath_manufacturer):

    browser = webdriver.Chrome('C:\Python\Parsing\Eurotec_UpDate\chromediver\chromedriver.exe')

    try:
        browser.get('https://eurotec.ua/edit')
        time.sleep(1)
        username_input = browser.find_element(By.NAME, 'login')
        username_input.clear()
        username_input.send_keys(auth_date.username)
        time.sleep(0)
        password_input = browser.find_element(By.NAME, 'pass')
        password_input.clear()
        password_input.send_keys(auth_date.password)
        time.sleep(0)
        password_input.send_keys(Keys.ENTER)
        time.sleep(2)
        items_menu = browser.find_element(By.XPATH, '//*[@id="header"]/div/div[2]/div[1]/a[3]').click()
        time.sleep(2)
        import_button = browser.find_element(By.ID, 'importBtn').click()
        time.sleep(2)
        file_input = browser.find_element(By.XPATH, '/html/body/div[7]/div[2]/div/div[2]/div/div[1]/div[4]/label/input').send_keys(fr'C:\Python\Parsing\Eurotec_UpDate\{file_csv}')
        time.sleep(5)
        select_manufacturer = browser.find_element(By.CLASS_NAME, 'supplier-select__trigger').click()
        time.sleep(2)
        select_manufacturer = browser.find_element(By.XPATH, xpath_manufacturer).click()
        time.sleep(2)
        import_button = browser.find_element(By.XPATH, '/html/body/main/section/div[1]/div[2]/div/span[3]/div/span').click()
        time.sleep(2)
        add_button = browser.find_element(By.XPATH, '/html/body/div[9]/div[2]/div/div[2]/div/div/ul/li[1]/div[2]/div/div[2]').click()
        time.sleep(1)
        add_button = browser.find_element(By.XPATH, '/html/body/div[10]/div/div[2]').click()
        time.sleep(1)
        out_button = browser.find_element(By.XPATH, '/html/body/div[9]/div[2]/div/div[2]/div/div/ul/li[3]/div[2]/div/div[2]').click()
        time.sleep(1)
        out_button = browser.find_element(By.XPATH, '/html/body/div[10]/div/div[5]').click()
        time.sleep(1)
        fin_button = browser.find_element(By.XPATH, '/html/body/div[9]/div[2]/div/div[3]/div/span[2]').click()
        time.sleep(30)
    except Exception as ex:
        print(ex)
        browser.close()
        browser.quit()

def main():
    starttime = datetime.datetime.now()
    teka()
    franke()
    bsh()
    mirs()
    smeg()
    renklod()

    diftime = datetime.datetime.now() - starttime
    print(diftime)

if __name__ == '__main__':
    main()